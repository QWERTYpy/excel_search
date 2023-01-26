import openpyxl  # Библиотека для работы с таблицами
ip_book = openpyxl.load_workbook("result2.xlsx")  # Открывает файл
print(ip_book.sheetnames)
worksheet = ip_book.active  # Делаем его активным
max_row = worksheet.max_row  # Получаем максимальное количество строк
max_col = worksheet.max_column  # Получаем максимальное количество столбцо
print(max_col,max_row)
tabel = worksheet.cell(row=1, column=1).value
print(tabel)
print(worksheet['A1:C1'])
print(worksheet['A1':'C1'])
aa=worksheet[1]
print([x.value for x in aa])
tabel = worksheet.cell(row=1, column=1, value="2")
tabel = worksheet.cell(row=1, column=1).value
print(tabel)
ip_book.save('result2.xlsx')