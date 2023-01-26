from tqdm import tqdm
import openpyxl  # Библиотека для работы с таблицами
import pandas as pd
ip_book = openpyxl.load_workbook("start_file.xlsx")  # Открывает файл
worksheet = ip_book.active  # Делаем его активным
max_row = worksheet.max_row  # Получаем максимальное количество строк
max_col = worksheet.max_column  # Получаем максимальное количество столбцов

crash_date = ['02-16', '02-17', '02-18', '02-19', '02-20', '03-16',
              '03-25', '03-30', '04-01', '04-02', '04-03', '04-09']
marker_list = []
marker_dict = {}
for row in tqdm(range(1, max_row+1), desc='Обработка файла :'):
    tabel = worksheet.cell(row=row, column=2).value
    date_in = worksheet.cell(row=row, column=3).value
    date_in = str(date_in)[5:10]

    date_out = worksheet.cell(row=row, column=4).value
    date_out = str(date_out)[5:10]

    if date_in in crash_date:  # or date_out in crash_date:
        if tabel in marker_dict.keys():
            marker_dict[tabel][crash_date.index(date_in)] += 1
        else:
            marker_dict[tabel] = [0 for _ in range(len(crash_date))]
            marker_dict[tabel][crash_date.index(date_in)] += 1
    if date_out in crash_date:
        if tabel in marker_dict.keys():
            marker_dict[tabel][crash_date.index(date_out)] += 1
        else:
            marker_dict[tabel] = [0 for _ in range(len(crash_date))]
            marker_dict[tabel][crash_date.index(date_out)] += 1

lost = []
lost_min = []
for row in tqdm(range(1, max_row+1), desc='Обработка файла :'):
    tabel = worksheet.cell(row=row, column=2).value
    name = worksheet.cell(row=row, column=1).value
    org = worksheet.cell(row=row, column=9).value
    if tabel in marker_dict.keys():
        continue
    if tabel not in lost_min:
        lost_min.append(tabel)
        lost.append([tabel, name, org])

for _ in lost:
    print(_)

dict_crash = {}
dict_crash_tmp = []
count = 0
for tabel in marker_dict:
    if all(marker_dict.get(tabel)):
        dict_crash_tmp.append([tabel]+marker_dict.get(tabel))

for i in range(len(dict_crash_tmp[0])):
    if i == 0:
        dict_crash['tabel'] = [dict_crash_tmp[j][i] for j in range(len(dict_crash_tmp))]
    else:
        dict_crash[crash_date[i-1]] = [dict_crash_tmp[j][i] for j in range(len(dict_crash_tmp))]
df = pd.DataFrame(dict_crash)
df.to_excel('end_file.xlsx', index=False)
