# import docx
# from PyPDF2 import PdfFileReader

import openpyxl
# ip_book = openpyxl.load_workbook("result.xlsx")  # Открывает файл
# worksheet = ip_book.active
# x=['a','b']
# doc = docx.Document('www.docx')
# print(doc.)
from docx2pdf import convert
# convert("www.docx")
# pdf_document = "www.pdf"
# with open(pdf_document, "rb") as filehandle:
#     pdf = PdfFileReader(filehandle)
#
#     info = pdf.getDocumentInfo()
#     pages = pdf.getNumPages()
#     # print("Количество страниц в документе: %i\n\n" % pages)
#     # print("Мета-описание: ", info)
#     for i in range(pages):
#         page = pdf.getPage(i)
#         # print("Стр.", i, " мета: ", page, "\n\nСодержание;\n")
#         list_date = page.extractText().split("\n")
#     list_date = [x.strip() for x in list_date if len(x) > 1]
#     print(list_date)

# ip_book = openpyxl.load_workbook("result.xlsx")  # Открывает файл
# worksheet = ip_book.active  # Делаем его активным
# max_row = worksheet.max_row  # Получаем максимальное количество строк
# max_col = worksheet.max_column  # Получаем максимальное количество столбцов
# print(max_col,max_row)
# worksheet.append(["1","2","3","4"])
# max_row = worksheet.max_row  # Получаем максимальное количество строк
# max_col = worksheet.max_column  # Получаем максимальное количество столбцов
# print(max_col,max_row)
# worksheet.append(["1","2","3","4"])
# max_row = worksheet.max_row  # Получаем максимальное количество строк
# max_col = worksheet.max_column  # Получаем максимальное количество столбцов
# print(max_col,max_row)
# ip_book.save("111.xlsx")
import os

folder_path = "doc23"
# for folder_name in os.listdir(folder_path):
#     print(folder_name)
ip_book = openpyxl.load_workbook("result.xlsx")  # Открывает файл
worksheet = ip_book.active  # Делаем его активным
for file_name in os.listdir(f"{folder_path}"):
    print(">",file_name)
    fio, *_ = file_name.split(" ")
    # print(fio)
    convert(f"{folder_path}/{file_name}")
